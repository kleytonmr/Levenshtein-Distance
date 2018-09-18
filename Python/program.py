from openpyxl import load_workbook 
import levenshtein

# test database
baseA = load_workbook('../bases/base_a.xlsx')
baseB = load_workbook('../bases/base_b.xlsx')

nameA = baseA['Sheet1']
nameB = baseB['Sheet1']

for line_A in nameA.rows:
    for cell_A in line_A:
        for line_B in nameB.rows:
            for cell_B in line_B:
                if levenshtein.compare(cell_A.value,cell_B.value) >= 85:
                    print("Levenshtein: {0}, {1}. Nível de semelhança: {2}% " .format(cell_A.value,cell_B.value,levenshtein.compare(cell_A.value,cell_B.value)))

        




